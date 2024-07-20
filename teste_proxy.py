import openpyxl
import requests

def verificar_proxy(proxy_info):
    """
    Verifica se uma proxy está funcionando ao tentar acessar o Google.

    Parâmetros:
        proxy_info (str): String contendo a proxy no formato 'host:port:login:password'.

    Retorna:
        bool: True se a proxy está funcionando, False caso contrário.
    """
    proxy_parts = proxy_info.split(':')
    proxy_host = proxy_parts[0]
    proxy_port = proxy_parts[1]
    proxy_login = proxy_parts[2]
    proxy_password = proxy_parts[3]

    proxy_dict = {
        "http": f"http://{proxy_login}:{proxy_password}@{proxy_host}:{proxy_port}",
        "https": f"http://{proxy_login}:{proxy_password}@{proxy_host}:{proxy_port}"
    }

    try:
        response = requests.get('https://www.google.com', proxies=proxy_dict, timeout=5)
        if response.status_code == 200:
            print(f"Proxy {proxy_info} está funcionando.")
            return True
    except Exception as e:
        print(f"Erro ao verificar proxy {proxy_info}: {e}")
    return False

def verificar_proxies_no_excel(file_path):
    """
    Verifica todas as proxies listadas no arquivo Excel e retorna uma lista das proxies que estão funcionando.

    Parâmetros:
        file_path (str): Caminho para o arquivo Excel.

    Retorna:
        list: Lista de proxies que estão funcionando.
    """
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    proxies_funcionando = []

    for row in range(2, sheet.max_row + 1):
        proxy_info = sheet.cell(row=row, column=2).value
        if verificar_proxy(proxy_info):
            proxies_funcionando.append(proxy_info)

    # Criar uma nova planilha para proxies funcionando
    new_sheet_name = "Proxies Funcionando"
    if new_sheet_name in workbook.sheetnames:
        new_sheet = workbook[new_sheet_name]
    else:
        new_sheet = workbook.create_sheet(new_sheet_name)

    # Escrever as proxies funcionando na nova planilha
    for index, proxy in enumerate(proxies_funcionando, start=1):
        new_sheet.cell(row=index, column=1, value=proxy)

    # Salvar o arquivo Excel
    workbook.save(file_path)
    workbook.close()

    print(f"Proxies funcionando foram salvos na planilha '{new_sheet_name}'.")

# Exemplo de uso
if __name__ == "__main__":
    file_path = './nomes.xlsx'  # Substitua pelo caminho do seu arquivo Excel
    verificar_proxies_no_excel(file_path)
