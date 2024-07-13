import openpyxl

def copy_all_columns_to_new_sheet(file_path, new_sheet_name):
    # Carregar o arquivo Excel
    wb = openpyxl.load_workbook(file_path)
    
    # Criar uma nova planilha
    if new_sheet_name in wb.sheetnames:
        new_sheet = wb[new_sheet_name]
    else:
        new_sheet = wb.create_sheet(new_sheet_name)

    current_row = 1  # Inicializar a linha atual para colar dados

    # Iterar por cada planilha no arquivo
    for sheet_name in wb.sheetnames:
        if sheet_name == new_sheet_name:
            continue  # Pular a nova planilha criada para evitar loop infinito

        sheet = wb[sheet_name]
        max_row = sheet.max_row
        max_col = sheet.max_column

        # Copiar dados da planilha atual para a nova planilha
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                new_sheet.cell(row=current_row, column=col, value=sheet.cell(row=row, column=col).value)
            current_row += 1  # Mover para a próxima linha na nova planilha

    # Salvar o arquivo Excel
    wb.save(file_path)
    print(f"Dados copiados para a nova planilha '{new_sheet_name}' com sucesso!")

# Exemplo de uso
file_path = './nomes.xlsx'  # Substitua pelo caminho do seu arquivo Excel
new_sheet_name = 'CUPIDO'
copy_all_columns_to_new_sheet(file_path, new_sheet_name)
