import tkinter as tk
from tkinter import ttk
import random
import json
import requests
import openpyxl

# Definição das variáveis e arrays de configuração
platforms_array = ["macos", "windows", "linux"]
vendors_array = ["Google Inc. (Intel)", "Google Inc. (NVIDIA)", "Google Inc. (AMD)", "Google Inc.",
                 "Google Inc. (Microsoft)", "Google Inc. (Unknown)", "Intel Inc.",
                 "Google Inc. (NVIDIA Corporation)"]
renderers_array = {
    "Google Inc. (Intel)": [
        "ANGLE (Intel, Intel(R) HD Graphics Family Direct3D11 vs_5_0 ps_5_0, D3D11)",
        "ANGLE (Intel, Intel(R) UHD Graphics Direct3D11 vs_5_0 ps_5_0, D3D11)"
    ],
    "Google Inc. (NVIDIA)": [
        "ANGLE (NVIDIA, NVIDIA GeForce GTX 1060 6GB Direct3D11 vs_5_0 ps_5_0, D3D11-30.0.14.7247)",
        "ANGLE (NVIDIA, NVIDIA GeForce GTX 1050 Ti Direct3D11 vs_5_0 ps_5_0, D3D11-27.21.14.6172)"
    ],
    "Google Inc. (AMD)": [
        "ANGLE (AMD, AMD Radeon(TM) Vega 10 Graphics Direct3D11 vs_5_0 ps_5_0, D3D11)",
        "ANGLE (AMD, AMD Radeon(TM) Graphics Direct3D11 vs_5_0 ps_5_0, D3D11-30.0.13044.0)"
    ],
    "Google Inc.": [
        "ANGLE (Intel(R) UHD Graphics Direct3D11 vs_5_0 ps_5_0)",
        "ANGLE (Intel(R) UHD Graphics 620 Direct3D11 vs_5_0 ps_5_0)"
    ],
    "Google Inc. (Microsoft)": [
        "ANGLE (Microsoft, Microsoft Basic Render Driver Direct3D11 vs_5_0 ps_5_0, D3D11-10.0.19041.546)",
        "ANGLE (Microsoft, Microsoft Basic Render Driver Direct3D11 vs_5_0 ps_5_0, D3D11-6.3.9600.16505)"
    ],
    "Google Inc. (Unknown)": [
        "ANGLE (Unknown, Qualcomm(R) Adreno(TM) 630 GPU Direct3D11 vs_5_0 ps_5_0, D3D11)",
        "ANGLE (Unknown, Qualcomm(R) Adreno(TM) 630 GPU Direct3D11 vs_5_0 ps_5_0, D3D11-25.18.10500.0)"
    ],
    "Intel Inc.": [
        "Intel Iris OpenGL Engine"
    ],
    "Google Inc. (NVIDIA Corporation)": [
        "ANGLE (NVIDIA Corporation, GeForce RTX 3070/PCIe/SSE2, OpenGL 4.5.0 NVIDIA 461.40)"
    ]
}
cpu = [2, 4, 6, 8, 12, 16]
memory = [2, 4, 8]

def authentication(token: str):
    return token

def get_new_user_agent(platform: str, session: requests.Session):
    version_number = random.randint(101, 107)
    options = {
        "url": f"https://anty-api.com/fingerprints/useragent?browser_type=anty&browser_version=107&platform={platform}"
    }
    response = session.get(options["url"]).json()
    if "data" in response:
        return response["data"]
    else:
        raise Exception("Can't get new user agent, something went wrong")

def create_multiple_profiles(session: requests.Session, num_profiles: int):
    workbook = openpyxl.load_workbook("./nomes.xlsx")
    sheet = workbook.active
    completed_sheet = workbook.create_sheet("Completed")
    profiles_created = []

    for i in range(num_profiles):
        current_vendor = random.choice(vendors_array)
        current_renderer = random.choice(renderers_array[current_vendor])
        current_platform = random.choice(platforms_array)
        current_user_agent = get_new_user_agent(platform=current_platform, session=session)
        proxy_info = sheet.cell(row=2, column=2).value
        proxy_parts = proxy_info.split(':')
        proxy_name = sheet.cell(row=2, column=1).value
        proxy_login = proxy_parts[2]
        proxy_password = proxy_parts[3]

        proxy_host = proxy_parts[0]
        proxy_port = proxy_parts[1]

        options = {
            "url": "https://anty-api.com/browser_profiles",
            "data": {
                "platform": current_platform,
                "browserType": "anty",
                "useragent": {
                    "mode": "manual",
                    "value": current_user_agent
                },
                "webrtc": {
                    "mode": "altered",
                    "ipAddress": None
                },
                "canvas": {
                    "mode": "real"
                },
                "webgl": {
                    "mode": "real"
                },
                "webglInfo": {
                    "mode": "manual",
                    "vendor": current_vendor,
                    "renderer": current_renderer
                },
                "geolocation": {
                    "mode": "real",
                    "latitude": None,
                    "longitude": None
                },
                "cpu": {
                    "mode": "manual",
                    "value": random.choice(cpu)
                },
                "memory": {
                    "mode": "manual",
                    "value": random.choice(memory)
                },
                "timezone": {
                    "mode": "auto",
                    "value": None
                },
                "locale": {
                    "mode": "auto",
                    "value": None
                },
                "name": proxy_name,
                "tags": [
                    ""
                ],
                "notes": {
                },
                "proxy": {
                    "name": proxy_name,
                    "host": proxy_host,
                    "port": proxy_port,
                    "type": "http",
                    "login": proxy_login,
                    "password": proxy_password
                },
                "statusId": 0,
                "doNotTrack": False
            }
        }

        try:
            response = session.post(options["url"], headers=session.headers, data=json.dumps(options["data"]))
            profile_id = get_last_browser_profile(session=session)
            profiles_created.append(profile_id['id'])
            print(f"Perfil criado com sucesso para {proxy_name}, ID: {profile_id['id']}")

            completed_sheet.append([
                sheet.cell(row=2, column=1).value,
                sheet.cell(row=2, column=2).value,
                sheet.cell(row=2, column=3).value,
                sheet.cell(row=2, column=4).value
            ])
            sheet.delete_rows(2)

        except Exception as e:
            print(f"Erro ao criar perfil para {proxy_name}: {str(e)}")

    workbook.save("./nomes.xlsx")
    workbook.close()
    session.close()
    return profiles_created

def get_last_browser_profile(session: requests.Session):
    options = {
        "url": "https://anty-api.com/browser_profiles"
    }
    response = session.get(options["url"]).json()
    if "data" in response:
        return response["data"][0]
    else:
        raise Exception("Can't get list of browser profiles")

def main(num_profiles):
    session = requests.session()
    auth_token = "eyJ0eXAiOiJKV1QiLCJhbGciOiJSUzI1NiJ9.eyJhdWQiOiIxIiwianRpIjoiODUwMjIyOGU0ZmMzYzMxNTA1ODUwNzdjNTZkM2Y5MGM2NWEzYWNiNGEwNDJlYTdlMTcwZDAyYTBlMGI1Yjc4MTAzNzg1MTYwNTYxMmJhYjkiLCJpYXQiOjE3MjE4OTY3NzguMTAwNjcsIm5iZiI6MTcyMTg5Njc3OC4xMDA2NzMsImV4cCI6MTc1MzQzMjc3OC4wODkzMjUsInN1YiI6IjI2OTAzMDQiLCJzY29wZXMiOltdfQ.pwbdn5lEHU5B9th8Q6et9dgocmM8P6ZwP72MkoQtkUr3HXE_qjcN2a4LMwk6Qgyr8RKaJt1raSnhTR9V97aczQhloetwDKG4iaLPWqaX-dVXkjW9P3Ph-oklQDIc6XMpqQB7XvEOrYh0qEp33u75ADHPHsL8sOfytcs1XCfE3KWcbHkISTF1CMHeROTT8BkZRh_Ic4TnvDZuwsZNkkkaOi521kTl_7dSo9582qwnspMWtK8tNvNhWVYp4PvTV1LaWsQsTJUGB454b97TGhkH0lpQF2JyKKPkw29dC08xTcudMAHPgEZR7zRQX5tMq74A3MnIL-tJ0Zy2kDtufaTN6Uv9ty27aaxKkvgyAYvBdDYFwPxg_fAotTurLtXlXV7Rkqx6sULmfzvZLsIbHbKSEpcIZIlNNzAl--4UgxcZTWIRzBXHPw6DejWaqgADYKQiKTv2A_95zl3L0GCuxvMVPYoba3b8-Ml2pEDOft9wwk5kERWPCTMN6Y_pG_b765IGnk0p-4Ib_TrlFAVByYO0wQBLjnyKiOMDonCnRgowwjQN13VMQu4ivnfjLst4e8NabQ27u-9Kly6Dpx51rg7ou-FyZKp5Set4-JEzM4SfbNotAv8zbSxnWoyZKc8L2wVakuv0DKLsXKFrYscA-UiwBmuq-iOUS3ZYA11uh56rAtE"
    session.headers.update({
        "Authorization": f"Bearer {authentication(token=auth_token)}",
        "Content-Type": "application/json"
    })

    profiles_created = create_multiple_profiles(session=session, num_profiles=num_profiles)
    return profiles_created

def start_creation():
    num_profiles = int(num_profiles_entry.get())
    created_profiles = main(num_profiles)
    display_profiles(created_profiles)

def display_profiles(profiles):
    for profile in profiles:
        profiles_listbox.insert(tk.END, profile)

# Interface Gráfica com tkinter
root = tk.Tk()
root.title("Gerador de Perfis de Navegador")

frame = ttk.Frame(root, padding="10")
frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

num_profiles_label = ttk.Label(frame, text="Número de Perfis:")
num_profiles_label.grid(row=0, column=0, sticky=tk.W)

num_profiles_entry = ttk.Entry(frame)
num_profiles_entry.grid(row=0, column=1, sticky=(tk.W, tk.E))

start_button = ttk.Button(frame, text="Iniciar Criação", command=start_creation)
start_button.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E))

profiles_listbox = tk.Listbox(frame, height=10, width=50)
profiles_listbox.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E))

root.mainloop()