import random
import json
import requests
import openpyxl

# Definição das variáveis e arrays de configuração
platforms_array = ["macos", "windows", "linux"]
vendors_array = ["Google Inc. (Intel)", "Google Inc. (NVIDIA)", "Google Inc. (AMD)", "Google Inc.",
                 "Google Inc. (Microsoft)", "Google Inc. (Unknown)", "Intel Inc.",
                 "Google Inc. (NVIDIA Corporation)"]
renderers_array = {
    "Google Inc. (Intel)": [
        "ANGLE (Intel, Intel(R) HD Graphics Family Direct3D11 vs_5_0 ps_5_0, D3D11)",
        "ANGLE (Intel, Intel(R) UHD Graphics Direct3D11 vs_5_0 ps_5_0, D3D11)"
    ],
    "Google Inc. (NVIDIA)": [
        "ANGLE (NVIDIA, NVIDIA GeForce GTX 1060 6GB Direct3D11 vs_5_0 ps_5_0, D3D11-30.0.14.7247)",
        "ANGLE (NVIDIA, NVIDIA GeForce GTX 1050 Ti Direct3D11 vs_5_0 ps_5_0, D3D11-27.21.14.6172)"
    ],
    "Google Inc. (AMD)": [
        "ANGLE (AMD, AMD Radeon(TM) Vega 10 Graphics Direct3D11 vs_5_0 ps_5_0, D3D11)",
        "ANGLE (AMD, AMD Radeon(TM) Graphics Direct3D11 vs_5_0 ps_5_0, D3D11-30.0.13044.0)"
    ],
    "Google Inc.": [
        "ANGLE (Intel(R) UHD Graphics Direct3D11 vs_5_0 ps_5_0)",
        "ANGLE (Intel(R) UHD Graphics 620 Direct3D11 vs_5_0 ps_5_0)"
    ],
    "Google Inc. (Microsoft)": [
        "ANGLE (Microsoft, Microsoft Basic Render Driver Direct3D11 vs_5_0 ps_5_0, D3D11-10.0.19041.546)",
        "ANGLE (Microsoft, Microsoft Basic Render Driver Direct3D11 vs_5_0 ps_5_0, D3D11-6.3.9600.16505)"
    ],
    "Google Inc. (Unknown)": [
        "ANGLE (Unknown, Qualcomm(R) Adreno(TM) 630 GPU Direct3D11 vs_5_0 ps_5_0, D3D11)",
        "ANGLE (Unknown, Qualcomm(R) Adreno(TM) 630 GPU Direct3D11 vs_5_0 ps_5_0, D3D11-25.18.10500.0)"
    ],
    "Intel Inc.": [
        "Intel Iris OpenGL Engine"
    ],
    "Google Inc. (NVIDIA Corporation)": [
        "ANGLE (NVIDIA Corporation, GeForce RTX 3070/PCIe/SSE2, OpenGL 4.5.0 NVIDIA 461.40)"
    ]
}
cpu = [2, 4, 6, 8, 12, 16]
memory = [2, 4, 8]

# Função para autenticar e obter o token
def authentication(token: str):
    # Simplesmente retorna o token recebido
    return token

# Função para obter um novo user agent para uma determinada plataforma
def get_new_user_agent(platform: str, session: requests.Session):
    version_number = random.randint(101, 107)
    options = {
        "url": f"https://anty-api.com/fingerprints/useragent?browser_type=anty&browser_version=107&platform={platform}"
    }
    response = session.get(options["url"]).json()
    if "data" in response:
        return response["data"]
    else:
        raise Exception("Can't get new user agent, something went wrong")

# Função para criar múltiplos perfis de navegador
def create_multiple_profiles(session: requests.Session, num_profiles: int):
    workbook = openpyxl.load_workbook("./nomes.xlsx")
    sheet = workbook.active
    completed_sheet = workbook.create_sheet("Completed")
    profiles_created = []

    for i in range(num_profiles):
        current_vendor = random.choice(vendors_array)
        current_renderer = random.choice(renderers_array[current_vendor])
        current_platform = random.choice(platforms_array)
        current_user_agent = get_new_user_agent(platform=current_platform, session=session)
        proxy_info = sheet.cell(row=2, column=2).value  # Assuming proxy info is in the second column, starting from row 2
        proxy_parts = proxy_info.split(':')
        proxy_name = sheet.cell(row=2, column=1).value  # Assuming names are in the first column, starting from row 2
        proxy_login = proxy_parts[2]  # Assuming login is the third element in proxy_parts
        proxy_password = proxy_parts[3]  # Assuming password is the fourth element in proxy_parts
        facebook_login = sheet.cell(row=2, column=3).value  # Assuming Facebook login is in the third column, starting from row 2
        facebook_password = sheet.cell(row=2, column=4).value  # Assuming Facebook password is in the fourth column, starting from row 2

        proxy_host = proxy_parts[0]
        proxy_port = proxy_parts[1]

        options = {
            "url": "https://anty-api.com/browser_profiles",
            "data": {
                "platform": current_platform,
                "browserType": "anty",
                "useragent": {
                    "mode": "manual",
                    "value": current_user_agent
                },
                "webrtc": {
                    "mode": "altered",
                    "ipAddress": None
                },
                "canvas": {
                    "mode": "real"
                },
                "webgl": {
                    "mode": "real"
                },
                "webglInfo": {
                    "mode": "manual",
                    "vendor": current_vendor,
                    "renderer": current_renderer
                },
                "geolocation": {
                    "mode": "real",
                    "latitude": None,
                    "longitude": None
                },
                "cpu": {
                    "mode": "manual",
                    "value": random.choice(cpu)
                },
                "memory": {
                    "mode": "manual",
                    "value": random.choice(memory)
                },
                "timezone": {
                    "mode": "auto",
                    "value": None
                },
                "locale": {
                    "mode": "auto",
                    "value": None
                },
                "name": proxy_name,
                "tags": [
                    ""
                ],
                "mainWebsite": "facebook",
                "notes": {
                    "username": facebook_login,
                    "password": facebook_password,
                    "content": None,
                    "color": "blue",
                    "style": "text",
                    "icon": None
                },
                "proxy": {
                    "name": f"http://{proxy_name}",
                    "host": proxy_host,
                    "port": proxy_port,
                    "type": "http",
                    "login": proxy_login,
                    "password": proxy_password
                },
                "statusId": 0,
                "doNotTrack": False
            }
        }

        try:
            response = session.post(options["url"], headers=session.headers, data=json.dumps(options["data"]))
            profile_id = get_last_browser_profile(session=session)
            profiles_created.append(profile_id['id'])
            print(f"Perfil criado com sucesso para {proxy_name}, ID: {profile_id['id']}")

            # Mover informações para a planilha "Completed"
            completed_sheet.append([
                sheet.cell(row=2, column=1).value,  # Nome do perfil
                sheet.cell(row=2, column=2).value,  # Proxy
                sheet.cell(row=2, column=3).value,  # Facebook login
                sheet.cell(row=2, column=4).value   # Facebook password
            ])

            # Excluir a linha original
            sheet.delete_rows(2)

        except Exception as e:
            print(f"Erro ao criar perfil para {proxy_name}: {str(e)}")

    workbook.save("./nomes.xlsx")
    workbook.close()
    session.close()
    return profiles_created

# Função para obter o último perfil de navegador criado
def get_last_browser_profile(session: requests.Session):
    options = {
        "url": "https://anty-api.com/browser_profiles"
    }
    response = session.get(options["url"]).json()
    if "data" in response:
        return response["data"][0]
    else:
        raise Exception("Can't get list of browser profiles")

# Função principal para executar o código
def main():
    session = requests.session()
    auth_token = "eyJ0eXAiOiJKV1QiLCJhbGciOiJSUzI1NiJ9.eyJhdWQiOiIxIiwianRpIjoiNWJkOGJjNmRjZWMyMWJiNDM5NzMyNmNkMTJiN2Y2YmQzNGYzZjZjZDFhZTU3ZDZkOWQwYzJhODkyNTliNmJkMDM3ZGI3NzgyYWZiYTA3ODkiLCJpYXQiOjE3MTk2MDIxNDEuNDcyMzM1LCJuYmYiOjE3MTk2MDIxNDEuNDcyMzM4LCJleHAiOjE3MjIxOTQxNDEuNDU3ODQ4LCJzdWIiOiIyNjkwMzA0Iiwic2NvcGVzIjpbXX0.Xxa2zDv0AsXOEU0frZNrAbRkOaoTEY4uuSIe308s1ZnarRV2FGtI6YEtQdRQE3YPlTs11ro6iiT8jzJhAiNUkmuEwAvo2ISNTINWQHdVFUsi4tscmmp-VwcmKAFhMBrUwpI7DYJ4JxP41fk2N33oxmmpRKC9huPkb_dMCaN_o01kt03SuhSY2kpQsW75R_Y3vOwHuG8jjEZ9nNMvx1p0JB0sLth7NhLMTsvcmuNuJch8P5YmWDBj3nLV4npUilnIJkw8e-b6vsGPHLB0p6TtyY-75pXUAQVKwCbJHtODTQPc0lG9PwWC-wltEVcwOFwQc-hp2T7w40YUEA9NGyv1PwKDXUPucf1NIYFpaHZlvAIWDwECuGt5Yc6swRgkayemSXbs4fn6TIEzYHVR5yFJTYk14La36wGrrbJiVXiR0NuNBMr8MovqAX12Cu6L9LHQHKmIZjqdvQVycBD_-zWe8_yiTFwGadHP6O5QU0VyccmA_RT09pMncawhfcTdXOM9T1olgVe-adbvpGQbaTs8Hz4bvxTYbV8YDQ9p4XRk2_DNjXvTxMWmiXPBZ09OF_OMlce8tlb62eNhx9kIXX8VXMdcOvpYgZR2rBU2DzhHE7N33Na-DGuKbh41NTUxQGMaVe2No11XBHH4LGsvJjcNVZDKyTmyqwla_dUYyxD2mIA"
    session.headers = {
        "Authorization": "Bearer " + authentication(auth_token),
        "Content-Type": "application/json"
    }
    profiles_created = create_multiple_profiles(
        session=session, num_profiles=2)
    print("Perfis de navegador criados:", profiles_created)

if __name__ == "__main__":
    main()
